import csv
from io import BytesIO
from io import StringIO
from datetime import date
from datetime import datetime
from datetime import time

from fastapi import APIRouter
from fastapi import Depends
from fastapi import HTTPException
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from Backend.database import get_db
from Backend.services.auth_service import require_permission
from Backend.services.reporte_service import reporte_general
from Backend.services.reporte_service import historial_operaciones
from Backend.schemas.extraccion_cuenta import ExtraccionCuentaCreate
from Backend.services.cuenta_service import (
    crear_extraccion_cuenta,
    listar_extracciones_cuenta,
    listar_saldos_cuenta
)


def _normalizar_fecha_desde(value: date | None):
    if value is None:
        return None
    return datetime.combine(
        value,
        time.min
    )


def _normalizar_fecha_hasta(value: date | None):
    if value is None:
        return None
    return datetime.combine(
        value,
        time.max
    )


router = APIRouter(
    prefix="/reportes",
    tags=["Reportes"],
    dependencies=[
        Depends(
            require_permission(
                "pedidos:gestionar"
            )
        )
    ]
)


@router.get(
    "/resumen"
)
def resumen_route(
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    estado: str | None = None,
    servicio: str | None = None,
    moneda_pago: str | None = None,
    operador_id: int | None = None,
    metodo_pago_id: int | None = None,
    cuenta_pago_id: int | None = None,
    db: Session = Depends(
        get_db
    )
):
    return reporte_general(
        db,
        fecha_desde=_normalizar_fecha_desde(fecha_desde),
        fecha_hasta=_normalizar_fecha_hasta(fecha_hasta),
        estado=estado,
        servicio=servicio,
        moneda_pago=moneda_pago,
        operador_id=operador_id,
        metodo_pago_id=metodo_pago_id,
        cuenta_pago_id=cuenta_pago_id
    )


@router.get(
    "/resumen.csv"
)
def resumen_csv_route(
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    estado: str | None = None,
    servicio: str | None = None,
    moneda_pago: str | None = None,
    operador_id: int | None = None,
    metodo_pago_id: int | None = None,
    cuenta_pago_id: int | None = None,
    db: Session = Depends(
        get_db
    )
):
    reporte = reporte_general(
        db,
        fecha_desde=_normalizar_fecha_desde(fecha_desde),
        fecha_hasta=_normalizar_fecha_hasta(fecha_hasta),
        estado=estado,
        servicio=servicio,
        moneda_pago=moneda_pago,
        operador_id=operador_id,
        metodo_pago_id=metodo_pago_id,
        cuenta_pago_id=cuenta_pago_id
    )

    output = StringIO()
    writer = csv.writer(
        output
    )
    writer.writerow(
        [
            "seccion",
            "clave",
            "cantidad",
            "monto_pago",
            "ganancia"
        ]
    )

    resumen = reporte[
        "resumen"
    ]
    writer.writerow(
        [
            "resumen",
            "total",
            resumen[
                "total_pedidos"
            ],
            resumen[
                "monto_pago_total"
            ],
            resumen[
                "ganancia_total"
            ]
        ]
    )

    for seccion in [
        "por_dia",
        "por_estado",
        "por_servicio",
        "por_moneda",
        "por_metodo_pago",
        "por_cuenta_pago",
        "por_operador"
    ]:
        for fila in reporte[
            seccion
        ]:
            writer.writerow(
                [
                    seccion,
                    fila[
                        "clave"
                    ],
                    fila[
                        "cantidad"
                    ],
                    fila[
                        "monto_pago"
                    ],
                    fila[
                        "ganancia"
                    ]
                ]
            )

    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=reporte_resumen.csv"
        }
    )


@router.get(
    "/operaciones.xlsx"
)
def operaciones_xlsx_route(
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
    estado: str | None = None,
    servicio: str | None = None,
    moneda_pago: str | None = None,
    operador_id: int | None = None,
    metodo_pago_id: int | None = None,
    cuenta_pago_id: int | None = None,
    db: Session = Depends(
        get_db
    )
):
    operaciones = historial_operaciones(
        db,
        fecha_desde=_normalizar_fecha_desde(fecha_desde),
        fecha_hasta=_normalizar_fecha_hasta(fecha_hasta),
        estado=estado,
        servicio=servicio,
        moneda_pago=moneda_pago,
        operador_id=operador_id,
        metodo_pago_id=metodo_pago_id,
        cuenta_pago_id=cuenta_pago_id
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Operaciones"
    encabezados = [
        "Dia",
        "Fecha",
        "Codigo",
        "Gestor",
        "Banco / Cuenta",
        "Moneda",
        "Monto pago",
        "Oferta / Tasa",
        "Transf. CUP",
        "USD",
        "Efectivo CUP",
        "Monto MLC",
        "Recarga",
        "Otros",
        "Ganancia",
        "Estado",
        "Observaciones",
    ]
    worksheet.append(encabezados)

    dias = [
        "lunes",
        "martes",
        "miercoles",
        "jueves",
        "viernes",
        "sabado",
        "domingo",
    ]
    for operacion in operaciones:
        fecha = operacion["fecha"]
        worksheet.append([
            dias[fecha.weekday()] if fecha else "",
            fecha,
            operacion["codigo"],
            operacion["gestor"],
            operacion["banco"],
            operacion["moneda"],
            operacion["monto_pago"],
            operacion["tasa"],
            operacion["transferencia_cup"],
            operacion["usd"],
            operacion["efectivo_cup"],
            operacion["mlc"],
            operacion["recarga"],
            operacion["otros"],
            operacion["ganancia"],
            operacion["estado"],
            operacion["observaciones"],
        ])

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(
            color="FFFFFF",
            bold=True
        )
        cell.alignment = Alignment(
            horizontal="center"
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 13
    worksheet.column_dimensions["B"].width = 19
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 20
    worksheet.column_dimensions["E"].width = 28
    worksheet.column_dimensions["F"].width = 12
    for column in range(7, 16):
        worksheet.column_dimensions[get_column_letter(column)].width = 16
    worksheet.column_dimensions["P"].width = 20
    worksheet.column_dimensions["Q"].width = 36

    for row in worksheet.iter_rows(
        min_row=2
    ):
        row[1].number_format = "dd/mm/yyyy hh:mm"
        for cell in row[6:15]:
            cell.number_format = '#,##0.00'

    output = BytesIO()
    workbook.save(output)
    filename = (
        "operaciones_"
        + (fecha_desde.isoformat() if fecha_desde else "inicio")
        + "_"
        + (fecha_hasta.isoformat() if fecha_hasta else "hoy")
        + ".xlsx"
    )

    return Response(
        content=output.getvalue(),
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


@router.get("/cuentas/saldos")
def saldos_cuenta_route(
    metodo_pago_id: int | None = None,
    cuenta_pago_id: int | None = None,
    db: Session = Depends(get_db)
):
    return listar_saldos_cuenta(
        db,
        metodo_pago_id=metodo_pago_id,
        cuenta_pago_id=cuenta_pago_id
    )


@router.get("/extracciones")
def extracciones_cuenta_route(
    cuenta_pago_id: int | None = None,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    return listar_extracciones_cuenta(
        db,
        cuenta_pago_id=cuenta_pago_id,
        limit=limit
    )


@router.post("/extracciones")
def crear_extraccion_cuenta_route(
    data: ExtraccionCuentaCreate,
    db: Session = Depends(get_db),
    operador = Depends(require_permission("pedidos:gestionar"))
):
    try:
        return crear_extraccion_cuenta(db, data, operador)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
